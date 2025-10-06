import openpyxl

def read_test_cases(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    test_cases = []
    for row in range(2, sheet.max_row + 1):
        tc = {
            "Test Case ID": sheet.cell(row, 1).value,
            "Module": sheet.cell(row, 2).value,
            "Test Case Title": sheet.cell(row, 3).value,
            "Priority": sheet.cell(row, 4).value,
            "Test Type": sheet.cell(row, 5).value,
            "Pre-Conditions": sheet.cell(row, 6).value,
            "Test Steps": sheet.cell(row, 7).value,
            "Expected Result": sheet.cell(row, 8).value,
        }
        test_cases.append(tc)
    return test_cases
